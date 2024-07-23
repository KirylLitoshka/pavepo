import asyncio
import re
import logging
from openpyxl import Workbook, load_workbook
from pyppeteer import launch
from bs4 import BeautifulSoup

logging.basicConfig(level=logging.INFO)


def get_data_urls():
    urls = []
    workbook = load_workbook('data.xlsx')
    sheet = workbook.active
    for row in sheet:
        urls.append(row[0].value)
    workbook.close()
    return urls


async def save_data(content):
    wb = Workbook()
    sheet = wb.active
    for row in content:
        for key in row.keys():
            sheet.append([
                key,
                row[key]['subs'],
                row[key]['socials'].get('twitter', ""),
                row[key]['socials'].get('facebook', ""),
                row[key]['socials'].get('instagram', ""),
                row[key]['socials'].get('youtube', ""),
                row[key]['socials'].get('discord', ""),
                row[key]['socials'].get('vk', ""),
                row[key]['socials'].get('boosty', ""),
                row[key]['socials'].get('tg', ""),
                row[key]['mail']
            ])
    wb.save("output.xlsx")


def parse_socials(socials_content):
    socials = {}
    if socials_content is None:
        return socials
    for social in socials_content:
        link = social.find('a').get('href', None)
        if link is None:
            continue
        if link.lower().startswith('https://twitter.com'):
            socials['twitter'] = link
        elif link.lower().startswith('https://www.facebook.com'):
            socials['facebook'] = link
        elif link.lower().startswith('https://www.instagram.com'):
            socials['instagram'] = link
        elif link.lower().startswith('https://www.youtube.com'):
            socials['youtube'] = link
        elif link.lower().startswith('https://discordapp.com'):
            socials['discord'] = link
        elif link.lower().startswith('https://vk.com'):
            socials['vk'] = link
        elif link.lower().startswith('https://boosty.to'):
            socials['boosty'] = link
        elif link.lower().startswith('https://t.me'):
            socials['tg'] = link
        else:
            continue
    return socials


def parse_email(channel_content):
    soup_string = str(channel_content)
    mail_pattern = re.compile(r"\b[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Z|a-z]{2,7}\b")
    matches = mail_pattern.findall(soup_string)
    if matches:
        return matches[0]
    return ""


async def get_page_content(url: str, semaphore: asyncio.Semaphore):
    async with semaphore:
        try:
            browser = await launch(headless=True)
            browser_page = await browser.newPage()
            await browser_page.goto(url + "/about", {'waitUntil': 'networkidle2'})
            content = await browser_page.content()
            await browser.close()
            return {url: content}
        except Exception:
            return {url: None}


async def get_urls_data(urls, limit):
    sem = asyncio.Semaphore(limit)
    tasks = []
    for url in urls:
        tasks.append(asyncio.create_task(get_page_content(url, sem)))
    result = await asyncio.gather(*tasks)
    return result


async def clean_urls_data(urls_data):
    for row in urls_data:
        for key in row.keys():
            row_content = row[key]
            if row_content is None:
                row[key] = {
                    'subs': '',
                    'socials': {},
                    'mail': ""
                }
                continue
            soup = BeautifulSoup(row_content, "html.parser")
            channel_info_content = soup.find("div", {"class": "channel-info-content"})
            socials_content = channel_info_content.find_all("div", {"class": "social-media-link"})
            row[key] = {}
            followers = channel_info_content.find("span", {"class": "iFvAnD"})
            row[key]['subs'] = followers.text if followers is not None else None
            row[key]['socials'] = parse_socials(socials_content)
            row[key]['mail'] = parse_email(channel_info_content)
    return urls_data


async def main():
    twitch_urls = get_data_urls()
    concurrent_request_limit = 5  # Количество одновременных запросов (не для слабых пк)
    urls_data = await get_urls_data(twitch_urls[1:40], concurrent_request_limit)
    urls_content = await clean_urls_data(urls_data)
    await save_data(urls_content)


if __name__ == "__main__":
    asyncio.get_event_loop().run_until_complete(main())
